#!/usr/bin/env python3
"""
Importar nómina abril 2026 a A&K Supabase — VERSIÓN DIRECTA
 Usa la SUPABASE_SERVICE_ROLE_KEY directamente (de .env.supabase).
 
 PRERREQUISITOS:
   1. Ejecutar nomina_ddl_ak.sql en Supabase Dashboard SQL Editor
   2. pip install openpyxl requests
 
 Uso:
   python3 scripts/import_nomina_ak_v2.py
"""

import openpyxl
import requests
import json
import sys
import os
from datetime import datetime
from typing import Optional

# ═══════════════════════════════════════════════════════
# CONFIG — lee de .env.supabase
# ═══════════════════════════════════════════════════════

def load_env():
    """Load Supabase config from .env.supabase"""
    env_path = os.path.join(os.path.dirname(__file__), '..', '.env.supabase')
    env = {}
    if os.path.exists(env_path):
        with open(env_path, 'r') as f:
            for line in f:
                line = line.strip()
                if '=' in line and not line.startswith('#'):
                    k, v = line.split('=', 1)
                    env[k.strip()] = v.strip()
    
    # Also try .env.local (now has correct keys)
    env_local_path = os.path.join(os.path.dirname(__file__), '..', '.env.local')
    if os.path.exists(env_local_path):
        with open(env_local_path, 'r') as f:
            for line in f:
                line = line.strip()
                if '=' in line and not line.startswith('#'):
                    k, v = line.split('=', 1)
                    if k.strip() not in env:  # Don't override .env.supabase
                        env[k.strip()] = v.strip()
    return env

env = load_env()

SUPABASE_URL = env.get('NEXT_PUBLIC_SUPABASE_URL', '').rstrip('/')
SERVICE_KEY = env.get('SUPABASE_SERVICE_ROLE_KEY', '')
EXCEL_PATH = '/home/mod/.hermes/cache/documents/doc_8319112453a1_Nomina atico mes de abril  2026.xlsx'

if not SUPABASE_URL or not SERVICE_KEY:
    print("ERROR: No se encontraron SUPABASE_URL o SERVICE_KEY en .env.supabase o .env.local")
    sys.exit(1)

HEADERS = {
    'apikey': SERVICE_KEY,
    'Authorization': f'Bearer {SERVICE_KEY}',
    'Content-Type': 'application/json',
    'Prefer': 'return=representation',
}

# ═══════════════════════════════════════════════════════
# HELPERS
# ═══════════════════════════════════════════════════════

def safe_float(val, default=0):
    if val is None: return default
    try: return round(float(val), 2)
    except (ValueError, TypeError): return default

def safe_int(val, default=0):
    if val is None: return default
    try: return int(float(val))
    except (ValueError, TypeError): return default

def api_get(table, params=None):
    """GET from Supabase REST API"""
    url = f'{SUPABASE_URL}/rest/v1/{table}'
    r = requests.get(url, headers=HEADERS, params=params, timeout=30)
    return r

def api_post(table, data):
    """POST (insert) to Supabase REST API"""
    url = f'{SUPABASE_URL}/rest/v1/{table}'
    r = requests.post(url, headers=HEADERS, json=data, timeout=30)
    return r

def api_patch(table, match_params, data):
    """PATCH (update) to Supabase REST API"""
    url = f'{SUPABASE_URL}/rest/v1/{table}'
    headers = {**HEADERS}
    for k, v in match_params.items():
        url += f'&{k}=eq.{v}' if '?' in url else f'?{k}=eq.{v}'
    r = requests.patch(url, headers=headers, json=data, timeout=30)
    return r

# ═══════════════════════════════════════════════════════
# 0. CHECK STATUS
# ═══════════════════════════════════════════════════════

print("=" * 60)
print("NÓMINA A&K — IMPORTAR ABRIL 2026")
print("=" * 60)

print("\n🔍 Verificando estado de tablas...")
tables_status = {}
for table in ['pos_nomina_staff', 'nomina_periodos', 'nomina_detalle', 'nomina_he_recargos', 'nomina_novedades', 'nomina_provisiones', 'nomina_propinas']:
    r = api_get(table, {'select': 'id', 'limit': '1'})
    if r.status_code == 200:
        count_range = r.headers.get('content-range', '')
        total = count_range.split('/')[1] if '/' in count_range else '?'
        tables_status[table] = {'exists': True, 'count': total}
        print(f"  ✓ {table}: existe ({total} registros)")
    else:
        tables_status[table] = {'exists': False}
        print(f"  ✗ {table}: NO EXISTE")

# Check if pos_nomina_staff has the new columns
r = api_get('pos_nomina_staff', {'select': '*', 'limit': '1'})
if r.status_code == 200 and r.json():
    columns = list(r.json()[0].keys())
    needed = ['salario', 'cargo', 'modalidad', 'sede', 'fecha_ingreso', 'aplica_propinas', 'auxilio_no_salarial']
    missing = [c for c in needed if c not in columns]
    if missing:
        print(f"\n  ⚠️ pos_nomina_staff falta columnas: {missing}")
        print("  → Ejecutar DDL primero: nomina_ddl_ak.sql en Dashboard SQL Editor")
    else:
        print(f"\n  ✓ pos_nomina_staff tiene todas las columnas nuevas")
else:
    print(f"\n  ✗ No se pudo leer pos_nomina_staff")

missing_tables = [t for t, s in tables_status.items() if not s['exists']]
if missing_tables:
    print(f"\n❌ FALTAN TABLAS: {missing_tables}")
    print("   Ejecutá nomina_ddl_ak.sql en el Dashboard SQL Editor de Supabase primero.")
    print(f"   Archivo: /mnt/f/attick-keller/web/scripts/nomina_ddl_ak.sql")
    print("\n   Después de ejecutar el DDL, corré este script de nuevo.")
    sys.exit(1)

# ═══════════════════════════════════════════════════════
# 1. GET EXISTING DATA
# ═══════════════════════════════════════════════════════

print("\n📖 Obteniendo datos de BD...")

# Get all staff
r = api_get('pos_nomina_staff', {'select': 'id,cedula,nombre_completo,salario,cargo,sede', 'order': 'nombre_completo'})
staff_list = r.json() if r.status_code == 200 else []
staff_by_cedula = {s['cedula']: s for s in staff_list}
print(f"  Staff en BD: {len(staff_list)} empleados")

# Get periodos
r = api_get('nomina_periodos', {'select': '*'})
periodos = r.json() if r.status_code == 200 else []
periodo_by_sede = {p['sede']: p['id'] for p in periodos}
print(f"  Periodos: {len(periodos)}")

# ═══════════════════════════════════════════════════════
# 2. PARSE EXCEL
# ═══════════════════════════════════════════════════════

print("\n📖 Leyendo archivo Excel...")
wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
ws1 = wb['NOMINA MES DE ABRIL 2026']

sections = {
    'BAR_Y_MESAS_C75': (7, 35),
    'COCINA_C75': (39, 58),
    'CALLE_85': (62, 70),
    'KINDER': (74, 80),
    'ADMIN': (83, 88),
}

def parse_nomina_section(ws, start_row, end_row, default_sede):
    employees = []
    for row_num in range(start_row, end_row + 1):
        cedula = ws.cell(row=row_num, column=3).value
        nombre = ws.cell(row=row_num, column=4).value
        
        if cedula is None and nombre is None: continue
        if isinstance(cedula, str) and cedula.strip().upper() == 'CEDULA': continue
        if isinstance(nombre, str) and nombre.strip().upper() in ('NOMBRE DEL TRABAJADOR', 'TOTAL DEVENGADO', 'TOTALES DE NOMINAS'): continue
        
        empleado = {
            'cedula': str(int(cedula)) if isinstance(cedula, (int, float)) else str(cedula),
            'tipo_doc': str(ws.cell(row=row_num, column=2).value or 'CC').strip(),
            'nombre': str(nombre or '').strip(),
            'modalidad': str(ws.cell(row=row_num, column=5).value or 'COMPLETO').strip(),
            'fecha_ingreso': ws.cell(row=row_num, column=6).value,
            'fecha_retiro': ws.cell(row=row_num, column=7).value,
            'cargo': str(ws.cell(row=row_num, column=8).value or '').strip(),
            'dias': safe_int(ws.cell(row=row_num, column=9).value),
            'salario_basico': safe_float(ws.cell(row=row_num, column=10).value),
            'pagos_acordados': safe_float(ws.cell(row=row_num, column=11).value),
            'aplica_propinas': str(ws.cell(row=row_num, column=12).value or 'SI').strip().upper() == 'SI',
            'aplica_propinas_dias': safe_int(ws.cell(row=row_num, column=13).value),
            'prop_c75_dias': safe_int(ws.cell(row=row_num, column=14).value, None),
            'prop_c85_dias': safe_int(ws.cell(row=row_num, column=15).value, None),
            'salario': safe_float(ws.cell(row=row_num, column=16).value),
            'aux_transporte': safe_float(ws.cell(row=row_num, column=17).value),
            'propinas_prom_75_85': safe_float(ws.cell(row=row_num, column=18).value),
            'propinas_prom_75': safe_float(ws.cell(row=row_num, column=19).value),
            'aux_no_salarial': safe_float(ws.cell(row=row_num, column=20).value),
            'recargos_he_rn_rd': safe_float(ws.cell(row=row_num, column=21).value),
            'propinas': safe_float(ws.cell(row=row_num, column=22).value),
            'total_devengado': safe_float(ws.cell(row=row_num, column=23).value),
            'salud': safe_float(ws.cell(row=row_num, column=24).value),
            'pension': safe_float(ws.cell(row=row_num, column=25).value),
            'pagos_realizados': safe_float(ws.cell(row=row_num, column=26).value),
            'prestamos_consumos': safe_float(ws.cell(row=row_num, column=27).value),
            'total_deducciones': safe_float(ws.cell(row=row_num, column=28).value),
            'neto_a_pagar': safe_float(ws.cell(row=row_num, column=29).value),
            'sede': default_sede,
        }
        
        if isinstance(empleado['fecha_ingreso'], datetime):
            empleado['fecha_ingreso'] = empleado['fecha_ingreso'].strftime('%Y-%m-%d')
        if isinstance(empleado['fecha_retiro'], datetime):
            empleado['fecha_retiro'] = empleado['fecha_retiro'].strftime('%Y-%m-%d')
        
        if empleado['nombre'] and empleado['cedula']:
            employees.append(empleado)
    return employees

sede_map = {
    'BAR_Y_MESAS_C75': 'C75',
    'COCINA_C75': 'C75',
    'CALLE_85': 'C85',
    'KINDER': 'KINDER',
    'ADMIN': 'ADMIN',
}

all_employees = {}
for section_name, (start, end) in sections.items():
    sede = sede_map[section_name]
    employees = parse_nomina_section(ws1, start, end, sede)
    for emp in employees:
        key = f"{emp['cedula']}_{sede}"
        if key not in all_employees:
            all_employees[key] = emp

print(f"  Empleados únicos: {len(all_employees)}")

# Parse HE/Recargos
ws2 = wb['HE - REC DOMINICALES ABRIL 2026']
he_records = []
current_sede = 'C75'

for row_num in range(2, ws2.max_row + 1):
    nombre = ws2.cell(row=row_num, column=2).value
    if nombre and isinstance(nombre, str):
        upper = nombre.upper().strip()
        if 'CALLE 85' in upper: current_sede = 'C85'; continue
        if 'KINDER' in upper: current_sede = 'KINDER'; continue
        if 'ADMIN' in upper: current_sede = 'ADMIN'; continue
        if 'TOTAL' in upper or 'SUBTOTAL' in upper: continue
        if 'EMPLEADO' in upper: continue
    if not nombre or not str(nombre).strip(): continue
    num = ws2.cell(row=row_num, column=1).value
    if num is None and ws2.cell(row=row_num, column=3).value is None: continue
    
    he_records.append({
        'nombre': str(nombre).strip(),
        'sede': current_sede,
        'hed_horas': safe_float(ws2.cell(row=row_num, column=4).value),
        'hed_valor': safe_float(ws2.cell(row=row_num, column=5).value),
        'hed_total': safe_float(ws2.cell(row=row_num, column=6).value),
        'hen_horas': safe_float(ws2.cell(row=row_num, column=7).value),
        'hen_valor': safe_float(ws2.cell(row=row_num, column=8).value),
        'hen_total': safe_float(ws2.cell(row=row_num, column=9).value),
        'rn_horas': safe_float(ws2.cell(row=row_num, column=13).value),
        'rn_valor_hora': safe_float(ws2.cell(row=row_num, column=14).value),
        'rn_total': safe_float(ws2.cell(row=row_num, column=15).value),
        'rd_diurno_horas': safe_float(ws2.cell(row=row_num, column=16).value),
        'rd_diurno_valor': safe_float(ws2.cell(row=row_num, column=17).value),
        'rd_diurno_total': safe_float(ws2.cell(row=row_num, column=18).value),
        'rd_nocturno_horas': safe_float(ws2.cell(row=row_num, column=19).value),
        'rd_nocturno_valor': safe_float(ws2.cell(row=row_num, column=20).value),
        'rd_nocturno_total': safe_float(ws2.cell(row=row_num, column=21).value),
        'hedd_horas': safe_float(ws2.cell(row=row_num, column=22).value),
        'hedd_valor': safe_float(ws2.cell(row=row_num, column=23).value),
        'hedd_total': safe_float(ws2.cell(row=row_num, column=24).value),
        'hddn_horas': safe_float(ws2.cell(row=row_num, column=25).value),
        'hddn_valor': safe_float(ws2.cell(row=row_num, column=26).value),
        'hddn_total': safe_float(ws2.cell(row=row_num, column=27).value),
        'total_recargos': safe_float(ws2.cell(row=row_num, column=28).value),
    })

print(f"  HE/Recargos: {len(he_records)}")

# Parse Novedades
novedades = []
ws_nov75 = wb['NOVEDADES CALLE 75 ABRIL 26']
in_novedades = False
in_ingresos = False

for row_num in range(1, ws_nov75.max_row + 1):
    col1 = ws_nov75.cell(row=row_num, column=1).value
    if isinstance(col1, str) and 'INGRESO' in col1.upper(): in_ingresos = True; in_novedades = False; continue
    if isinstance(col1, str) and 'NOVEDADES' in col1.upper(): in_novedades = True; in_ingresos = False; continue
    if isinstance(col1, str) and 'RETIRO' in col1.upper(): in_novedades = False; in_ingresos = False; continue
    if isinstance(col1, str) and col1.strip().upper() == 'N°': continue
    
    if (in_novedades or in_ingresos) and col1 and isinstance(col1, (int, float)):
        cedula = ws_nov75.cell(row=row_num, column=2).value
        nombre = ws_nov75.cell(row=row_num, column=3).value or ''
        tipo = 'INGRESO' if in_ingresos else str(ws_nov75.cell(row=row_num, column=4).value or '').strip().upper()
        obs = str(ws_nov75.cell(row=row_num, column=5).value or '').strip() if in_novedades else str(ws_nov75.cell(row=row_num, column=4).value or '')
        dias = ws_nov75.cell(row=row_num, column=6 + (0 if in_novedades else 1)).value
        
        novedades.append({
            'cedula': str(int(cedula)) if isinstance(cedula, (int, float)) else str(cedula),
            'nombre': str(nombre).strip(),
            'tipo': tipo,
            'observacion': obs,
            'dias': str(dias) if dias else None,
            'sede': 'C75',
        })

print(f"  Novedades: {len(novedades)}")

# Parse Provisiones
ws_prov = wb['PROVISIONES ABRIL 2026']
provisiones = []
current_sede = 'C75'

for row_num in range(7, ws_prov.max_row + 1):
    col4 = ws_prov.cell(row=row_num, column=4).value
    if isinstance(col4, str):
        u = col4.strip().upper()
        if 'BAR Y MESAS' in u: current_sede = 'C75'; continue
        elif 'COCINA' in u: current_sede = 'C75'; continue
        elif 'CALLE 85' in u or 'PERSONAL CALLE 85' in u: current_sede = 'C85'; continue
        elif 'KINDER' in u: current_sede = 'KINDER'; continue
        elif 'ADMINISTRATIVO' in u: current_sede = 'ADMIN'; continue
        elif 'TOTAL' in u: continue
    
    col3 = ws_prov.cell(row=row_num, column=3).value
    if col3 is None or (isinstance(col3, str) and col3.strip().upper() == 'CEDULA'): continue
    nombre = ws_prov.cell(row=row_num, column=4).value
    if not nombre or not isinstance(nombre, str) or 'TOTAL' in nombre.upper(): continue
    
    cedula = col3
    provisiones.append({
        'cedula': str(int(cedula)) if isinstance(cedula, (int, float)) else str(cedula),
        'nombre': str(nombre).strip(),
        'sede': current_sede,
        'provisiones_salud': safe_float(ws_prov.cell(row=row_num, column=9).value),
        'provisiones_sociales': safe_float(ws_prov.cell(row=row_num, column=10).value),
        'base_vacaciones': safe_float(ws_prov.cell(row=row_num, column=11).value),
        'salud_empleado': safe_float(ws_prov.cell(row=row_num, column=12).value),
        'pension_empleado': safe_float(ws_prov.cell(row=row_num, column=13).value),
        'pension_empleador': safe_float(ws_prov.cell(row=row_num, column=14).value),
        'arl_empleador': safe_float(ws_prov.cell(row=row_num, column=15).value),
        'caja_empleador': safe_float(ws_prov.cell(row=row_num, column=16).value),
        'cesantias_empleador': safe_float(ws_prov.cell(row=row_num, column=17).value),
        'prima_empleador': safe_float(ws_prov.cell(row=row_num, column=18).value),
        'vacaciones_empleador': safe_float(ws_prov.cell(row=row_num, column=19).value),
        'intereses_cesantias_empleador': safe_float(ws_prov.cell(row=row_num, column=20).value),
    })

print(f"  Provisiones: {len(provisiones)}")

# ═══════════════════════════════════════════════════════
# 3. UPSERT EMPLOYEES
# ═══════════════════════════════════════════════════════

print("\n📝 Upserting empleados...")
upserted = 0
errors = 0

for key, emp in all_employees.items():
    cedula = emp['cedula']
    
    # Check if employee exists
    if cedula in staff_by_cedula:
        # Update existing employee with new fields
        staff_id = staff_by_cedula[cedula]['id']
        update_data = {
            'nombre_completo': emp['nombre'],
            'salario': emp['salario_basico'],
            'cargo': emp['cargo'],
            'modalidad': emp['modalidad'],
            'sede': emp['sede'],
            'aplica_propinas': emp['aplica_propinas'],
            'es_medio_tiempo': 'MEDIO' in emp['modalidad'].upper(),
            'auxilio_no_salarial': emp['aux_no_salarial'],
        }
        if emp.get('fecha_ingreso'):
            update_data['fecha_ingreso'] = emp['fecha_ingreso']
        
        r = api_patch('pos_nomina_staff', {'id': f'eq.{staff_id}'}, update_data)
        if r.status_code in (200, 204):
            upserted += 1
        else:
            print(f"  ✗ Error updating {cedula} ({emp['nombre']}): {r.text[:100]}")
            errors += 1
    else:
        # Insert new employee
        insert_data = {
            'cedula': cedula,
            'nombre_completo': emp['nombre'],
            'salario': emp['salario_basico'],
            'cargo': emp['cargo'],
            'modalidad': emp['modalidad'],
            'sede': emp['sede'],
            'aplica_propinas': emp['aplica_propinas'],
            'es_medio_tiempo': 'MEDIO' in emp['modalidad'].upper(),
            'auxilio_no_salarial': emp['aux_no_salarial'],
        }
        if emp.get('fecha_ingreso'):
            insert_data['fecha_ingreso'] = emp['fecha_ingreso']
        
        r = api_post('pos_nomina_staff', insert_data)
        if r.status_code in (200, 201):
            new_staff = r.json()
            new_id = new_staff[0]['id'] if isinstance(new_staff, list) else new_staff['id']
            staff_by_cedula[cedula] = {'id': new_id, 'cedula': cedula, 'nombre_completo': emp['nombre']}
            upserted += 1
        else:
            print(f"  ✗ Error inserting {cedula} ({emp['nombre']}): {r.text[:100]}")
            errors += 1

print(f"  Upserted: {upserted}, Errors: {errors}")

# Refresh staff_id_map
staff_id_map = {s['cedula']: s['id'] for s in staff_by_cedula.values()}

# ═══════════════════════════════════════════════════════
# 4. INSERT NOMINA_DETALLE
# ═══════════════════════════════════════════════════════

print("\n📊 Insertando nómina detalle...")
detalle_records = []
missing_staff = []

for key, emp in all_employees.items():
    cedula = emp['cedula']
    sede = emp['sede']
    periodo_id = periodo_by_sede.get(sede)
    staff_id = staff_id_map.get(cedula)
    
    if not periodo_id:
        print(f"  ⚠️ No periodo para sede {sede}")
        continue
    if not staff_id:
        missing_staff.append(f"{cedula} ({emp['nombre']})")
        continue
    
    detalle_records.append({
        'periodo_id': periodo_id,
        'staff_id': staff_id,
        'sede': sede,
        'dias_laborados': emp['dias'],
        'salario_basico': emp['salario_basico'],
        'salario_devengado': emp['salario'],
        'auxilio_transporte': emp['aux_transporte'],
        'propinas_prometido_75_85': emp['propinas_prom_75_85'],
        'propinas_prometido_75': emp['propinas_prom_75'],
        'auxilio_no_salarial': emp['aux_no_salarial'],
        'recargos_he_rn_rd': emp['recargos_he_rn_rd'],
        'propinas': emp['propinas'],
        'total_devengado': emp['total_devengado'],
        'salud_empleado': emp['salud'],
        'pension_empleado': emp['pension'],
        'pagos_realizados': emp['pagos_realizados'],
        'prestamos_consumos': emp['prestamos_consumos'],
        'total_deducciones': emp['total_deducciones'],
        'neto_a_pagar': emp['neto_a_pagar'],
    })

if missing_staff:
    print(f"  ⚠️ Sin staff_id: {missing_staff[:5]}")

# Insert in batches of 20
batch_size = 20
inserted_detalle = 0
for i in range(0, len(detalle_records), batch_size):
    batch = detalle_records[i:i+batch_size]
    r = api_post('nomina_detalle', batch)
    if r.status_code in (200, 201):
        inserted_detalle += len(batch)
    else:
        print(f"  ✗ Error inserting detalle batch {i//batch_size+1}: {r.text[:200]}")
        # Try one by one
        for rec in batch:
            r = api_post('nomina_detalle', rec)
            if r.status_code in (200, 201):
                inserted_detalle += 1
            else:
                print(f"    ✗ {rec['sede']}/{rec['staff_id']}: {r.text[:100]}")

print(f"  Inserted: {inserted_detalle}/{len(detalle_records)}")

# ═══════════════════════════════════════════════════════
# 5. INSERT HE/RECARGOS
# ═══════════════════════════════════════════════════════

print("\n⏱️ Insertando HE/Recargos...")
name_to_cedula = {}
for emp in all_employees.values():
    name_to_cedula[emp['nombre'].upper().strip()] = emp['cedula']
    parts = emp['nombre'].split()
    if len(parts) >= 2:
        name_to_cedula[parts[-1].upper()] = emp['cedula']

he_records_db = []
for rec in he_records:
    nombre_upper = rec['nombre'].upper().strip()
    cedula = None
    
    # Try exact match
    if nombre_upper in name_to_cedula:
        cedula = name_to_cedula[nombre_upper]
    else:
        # Try partial match
        for nk, nc in name_to_cedula.items():
            if nk in nombre_upper or nombre_upper in nk:
                cedula = nc
                break
        if not cedula:
            for emp in all_employees.values():
                if emp['nombre'].split()[-1].upper() in nombre_upper:
                    cedula = emp['cedula']
                    break
    
    staff_id = staff_id_map.get(cedula) if cedula else None
    if not staff_id:
        print(f"  ⚠️ Sin match HE para: {rec['nombre']}")
        continue
    
    sede = rec['sede']
    periodo_id = periodo_by_sede.get(sede)
    if not periodo_id: continue
    
    he_records_db.append({
        'periodo_id': periodo_id,
        'staff_id': staff_id,
        'sede': sede,
        **{k: v for k, v in rec.items() if k not in ('nombre', 'sede')},
    })

inserted_he = 0
for i in range(0, len(he_records_db), batch_size):
    batch = he_records_db[i:i+batch_size]
    r = api_post('nomina_he_recargos', batch)
    if r.status_code in (200, 201):
        inserted_he += len(batch)
    else:
        print(f"  ✗ Error HE batch: {r.text[:200]}")
        for rec in batch:
            r = api_post('nomina_he_recargos', rec)
            if r.status_code in (200, 201):
                inserted_he += 1

print(f"  Inserted: {inserted_he}/{len(he_records_db)}")

# ═══════════════════════════════════════════════════════
# 6. INSERT NOVEDADES
# ═══════════════════════════════════════════════════════

print("\n📋 Insertando novedades...")
novedades_db = []
for nov in novedades:
    staff_id = staff_id_map.get(nov['cedula'])
    if not staff_id:
        print(f"  ⚠️ No staff_id para novedad de {nov['cedula']}")
        continue
    sede = nov['sede']
    periodo_id = periodo_by_sede.get(sede)
    if not periodo_id: continue
    
    novedades_db.append({
        'periodo_id': periodo_id,
        'staff_id': staff_id,
        'sede': sede,
        'tipo': nov['tipo'],
        'observacion': nov.get('observacion', ''),
        'dias': nov.get('dias'),
    })

r = api_post('nomina_novedades', novedades_db)
if r.status_code in (200, 201):
    print(f"  Inserted: {len(novedades_db)}")
else:
    print(f"  Error batch: {r.text[:200]}")
    inserted = 0
    for nov in novedades_db:
        r = api_post('nomina_novedades', nov)
        if r.status_code in (200, 201): inserted += 1
    print(f"  Inserted one-by-one: {inserted}/{len(novedades_db)}")

# ═══════════════════════════════════════════════════════
# 7. INSERT PROVISIONES
# ═══════════════════════════════════════════════════════

print("\n💰 Insertando provisiones...")
provisiones_db = []
for prov in provisiones:
    cedula = prov['cedula']
    staff_id = staff_id_map.get(cedula)
    if not staff_id:
        nombre_upper = prov['nombre'].upper().strip()
        for emp in all_employees.values():
            if emp['nombre'].upper().strip() == nombre_upper or emp['nombre'].split()[-1].upper() in nombre_upper:
                staff_id = staff_id_map.get(emp['cedula'])
                break
    if not staff_id:
        print(f"  ⚠️ No staff_id para prov de {cedula} ({prov['nombre']})")
        continue
    
    sede = prov['sede']
    periodo_id = periodo_by_sede.get(sede)
    if not periodo_id: continue
    
    total_prov_emp = (prov['pension_empleador'] + prov['arl_empleador'] + prov['caja_empleador'] +
                      prov['cesantias_empleador'] + prov['prima_empleador'] + prov['vacaciones_empleador'] +
                      prov['intereses_cesantias_empleador'])
    
    provisiones_db.append({
        'periodo_id': periodo_id,
        'staff_id': staff_id,
        'sede': sede,
        'provisiones_salud': prov['provisiones_salud'],
        'provisiones_sociales': prov['provisiones_sociales'],
        'base_vacaciones': prov['base_vacaciones'],
        'salud_empleado': prov['salud_empleado'],
        'pension_empleado': prov['pension_empleado'],
        'pension_empleador': prov['pension_empleador'],
        'arl_empleador': prov['arl_empleador'],
        'caja_empleador': prov['caja_empleador'],
        'cesantias_empleador': prov['cesantias_empleador'],
        'prima_empleador': prov['prima_empleador'],
        'vacaciones_empleador': prov['vacaciones_empleador'],
        'intereses_cesantias_empleador': prov['intereses_cesantias_empleador'],
        'total_provision_empleador': round(total_prov_emp, 2),
    })

inserted_prov = 0
for i in range(0, len(provisiones_db), batch_size):
    batch = provisiones_db[i:i+batch_size]
    r = api_post('nomina_provisiones', batch)
    if r.status_code in (200, 201):
        inserted_prov += len(batch)
    else:
        print(f"  ✗ Error prov batch: {r.text[:200]}")
        for rec in batch:
            r = api_post('nomina_provisiones', rec)
            if r.status_code in (200, 201): inserted_prov += 1

print(f"  Inserted: {inserted_prov}/{len(provisiones_db)}")

# ═══════════════════════════════════════════════════════
# 8. INSERT PROPINAS
# ═══════════════════════════════════════════════════════

print("\n💵 Insertando propinas...")
propinas_data = [
    {'sede': 'C75', 'total_propinas_ventas': 60789692, 'prometidos_100_pct': 0, 'propina_para_rep': 60789692, 'dias_laborados_total': 1059, 'valor_dia_propina': 57402.92},
    {'sede': 'C85', 'total_propinas_ventas': 2827742, 'prometidos_100_pct': 0, 'propina_para_rep': 2827742, 'dias_laborados_total': 105, 'valor_dia_propina': 26930.88},
    {'sede': 'KINDER', 'total_propinas_ventas': 1969675, 'prometidos_100_pct': 0, 'propina_para_rep': 1969675, 'dias_laborados_total': 90, 'valor_dia_propina': 21885.28},
]

propinas_records = []
for p in propinas_data:
    sede = p['sede']
    periodo_id = periodo_by_sede.get(sede)
    if not periodo_id:
        print(f"  ⚠️ No periodo para propinas {sede}")
        continue
    propinas_records.append({**p, 'periodo_id': periodo_id})

inserted_prop = 0
for rec in propinas_records:
    r = api_post('nomina_propinas', rec)
    if r.status_code in (200, 201):
        inserted_prop += 1
    else:
        print(f"  ✗ Error propinas {rec['sede']}: {r.text[:200]}")

print(f"  Inserted: {inserted_prop}/{len(propinas_records)}")

# ═══════════════════════════════════════════════════════
# 9. UPDATE PERIODO TOTALS
# ═══════════════════════════════════════════════════════

print("\n📊 Actualizando totales de periodos...")
totals_by_sede = {}
for emp in all_employees.values():
    sede = emp['sede']
    if sede not in totals_by_sede:
        totals_by_sede[sede] = {'devengado': 0, 'deducciones': 0, 'neto': 0}
    totals_by_sede[sede]['devengado'] += emp['total_devengado']
    totals_by_sede[sede]['deducciones'] += emp['total_deducciones']
    totals_by_sede[sede]['neto'] += emp['neto_a_pagar']

for sede, totals in totals_by_sede.items():
    periodo_id = periodo_by_sede.get(sede)
    if not periodo_id: continue
    r = api_patch('nomina_periodos', {'id': f'eq.{periodo_id}'}, {
        'total_devengado': round(totals['devengado'], 2),
        'total_deducciones': round(totals['deducciones'], 2),
        'total_neto': round(totals['neto'], 2),
        'estado': 'CERRADO',
    })
    status = '✓' if r.status_code in (200, 204) else f'✗ {r.status_code}'
    print(f"  {sede}: Dev=${totals['devengado']:,.0f} Ded=${totals['deducciones']:,.0f} Neto=${totals['neto']:,.0f} {status}")

# ═══════════════════════════════════════════════════════
# RESUMEN
# ═══════════════════════════════════════════════════════

print("\n" + "=" * 60)
print("✅ IMPORTACIÓN COMPLETADA")
print("=" * 60)
print(f"  Empleados upserted: {upserted}")
print(f"  Detalle nómina: {inserted_detalle} registros")
print(f"  HE/Recargos: {inserted_he} registros")
print(f"  Novedades: {len(novedades_db)} registros")
print(f"  Provisiones: {inserted_prov} registros")
print(f"  Propinas: {inserted_prop} registros")
print("=" * 60)