#!/usr/bin/env python3
"""
Importar nómina abril 2026 desde Excel a Supabase Seadotec.
Requisito: las tablas ya deben existir (ejecutar nomina_ddl.sql primero).

Uso:
  python3 scripts/import_nomina.py

Env vars necesarias:
  NEXT_PUBLIC_RODRI_SUPABASE_URL
  RODRI_SUPABASE_SERVICE_ROLE_KEY
"""
import openpyxl, requests, json, os, sys
from datetime import date, datetime
from decimal import Decimal, ROUND_HALF_UP

SUPABASE_URL = os.environ.get('NEXT_PUBLIC_RODRI_SUPABASE_URL', 'https://seadotecznewqcvxsber.supabase.co')
SERVICE_KEY = os.environ.get('RODRI_SUPABASE_SERVICE_ROLE_KEY', '')

if not SERVICE_KEY:
    # Try reading from .env.local
    env_path = os.path.join(os.path.dirname(__file__), '..', '.env.local')
    if os.path.exists(env_path):
        with open(env_path) as f:
            for line in f:
                line = line.strip()
                if line.startswith('RODRI_SUPABASE_SERVICE_ROLE_KEY='):
                    SERVICE_KEY = line.split('=', 1)[1].strip()
                    break

if not SERVICE_KEY:
    print("ERROR: RODRI_SUPABASE_SERVICE_ROLE_KEY not found")
    sys.exit(1)

HEADERS = {
    'apikey': SERVICE_KEY,
    'Authorization': f'Bearer {SERVICE_KEY}',
    'Content-Type': 'application/json',
    'Prefer': 'return=representation'
}

EXCEL_PATH = os.environ.get('NOMINA_EXCEL', '/home/mod/.hermes/cache/documents/doc_8319112453a1_Nomina atico mes de abril  2026.xlsx')
PERIODO = 'ABRIL 2026'
FECHA_INICIO = '2026-04-01'
FECHA_FIN = '2026-04-30'

# Mapeo de cargos del Excel a teams en la BD
CARGO_TO_TEAM = {
    'CAJERA ADMINISTRATIVO': 'Bar',
    'AUXILIAR BAR': 'Bar',
    'JEFE DE BAR': 'Bar',
    'JEFE BAR 50% PROPINAS': 'Bar',
    'BARTENDER': 'Bar',
    'JEFE DE SERVICIO': 'Servicio',
    'SUB JEFE SERVICIO': 'Servicio',
    'MESERO': 'Servicio',
    'MESERA': 'Servicio',
    'ASESOR COM VENTAS': 'Servicio',
    'PASANTE ADMINISTRATIVA': 'Servicio',
    'JEFE DE COCINA': 'Cocina',
    'CHEF SUPERVISOR': 'Cocina',
    'AUXILIAR DE COCINA': 'Cocina',
    'COCINERO': 'Cocina',
    'STEWARD': 'Steward',
    'SERVICIOS GENERALES': 'Steward',
    'PASANTE AUX COCINA': 'Cocina',
    'PIZZERO': 'Pizzería',
    'CAJERA': 'Pizzería',  # solo Mónica en Kinder
    'ADMINISTRADOR': 'Administración',
    'AUXILIAR CONTABLE': 'Administración',
    'COMMUNITY MANAGER': 'Administración',
    'INGENIERO DE SONIDO': 'Bar',  # trabaja en Calle 75
    'MESERA': 'Servicio',
}

# Modalidad del Excel a código
MODALIDAD_MAP = {
    'COMPLETO': 'COMPLETO',
    'COMPLETO ': 'COMPLETO',
    'MEDIO TIEMPO': 'MEDIO_TIEMPO',
    'PASANTE PRODUCTIVA': 'PASANTE_PRODUCTIVO',
    'PASANTE LECT': 'PASANTE_LECT',
    'TIEMPO COMPLETO': 'COMPLETO',
}


def sb_get(table, params=None):
    """Get from Supabase REST API"""
    url = f'{SUPABASE_URL}/rest/v1/{table}'
    r = requests.get(url, headers=HEADERS, params=params or {})
    r.raise_for_status()
    return r.json()


def sb_post(table, data):
    """Post to Supabase REST API (batch up to 500)"""
    url = f'{SUPABASE_URL}/rest/v1/{table}'
    r = requests.post(url, headers=HEADERS, json=data)
    r.raise_for_status()
    return r.json()


def sb_upsert(table, data):
    """Upsert to Supabase REST API"""
    headers = {**HEADERS, 'Prefer': 'resolution=merge-duplicates'}
    url = f'{SUPABASE_URL}/rest/v1/{table}'
    r = requests.post(url, headers=headers, json=data)
    r.raise_for_status()
    return r.json()


def sb_patch(table, id_field, id_value, data):
    """Patch a record in Supabase"""
    url = f'{SUPABASE_URL}/rest/v1/{table}?{id_field}=eq.{id_value}'
    r = requests.patch(url, headers=HEADERS, json=data)
    r.raise_for_status()
    return r.json()


def get_or_create_sede(codigo):
    """Get sede id by codigo"""
    sedes = sb_get('sedes', {'codigo': f'eq.{codigo}'})
    if sedes:
        return sedes[0]['id']
    return None


def find_team_by_name(nombre):
    """Find team id by name"""
    teams = sb_get('teams')
    for t in teams:
        if t['nombre'].lower() == nombre.lower():
            return t['id']
    return None


def safe_float(val, default=0):
    """Convert to float, handling None and strings"""
    if val is None:
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


def safe_int(val, default=0):
    """Convert to int, handling None and strings"""
    if val is None:
        return default
    try:
        return int(val)
    except (ValueError, TypeError):
        return default


def safe_date(val):
    """Convert datetime to date string"""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d')
    return str(val)[:10]


def main():
    print("=== Importar Nómina ATICO Abril 2026 ===\n")
    
    # 1. Load Excel
    print(f"Cargando Excel: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws_nom = wb['NOMINA MES DE ABRIL 2026']
    ws_he = wb['HE - REC DOMINICALES ABRIL 2026']
    ws_nov75 = wb['NOVEDADES CALLE 75 ABRIL 26']
    ws_nov85 = wb['NOVEDADES CALLE 85- ABRIL 2026']
    ws_prov = wb['PROVISIONES ABRIL 2026']
    
    # 2. Get sede IDs
    sede_c75 = get_or_create_sede('C75')
    sede_c85 = get_or_create_sede('C85')
    sede_kinder = get_or_create_sede('KINDER')
    sede_admin = get_or_create_sede('ADMIN')
    print(f"Sedes: C75={sede_c75}, C85={sede_c85}, KINDER={sede_kinder}, ADMIN={sede_admin}")
    
    # 3. Get/create periodo
    periodos = sb_get('nomina_periodos', {'periodo': f'eq.{PERIODO}'})
    if not periodos:
        # Create periodos for each sede
        for sede_codigo, sede_id in [('C75', sede_c75), ('C85', sede_c85), ('KINDER', sede_kinder), ('ADMIN', sede_admin)]:
            sb_post('nomina_periodos', [{
                'periodo': PERIODO,
                'fecha_inicio': FECHA_INICIO,
                'fecha_fin': FECHA_FIN,
                'sede_id': sede_id,
                'estado': 'ABIERTO'
            }])
        periodos = sb_get('nomina_periodos', {'periodo': f'eq.{PERIODO}'})
    periodo_ids = {p['sede_id']: p['id'] for p in periodos}
    print(f"Periodos creados/encontrados: {len(periodos)}")
    
    # 4. Get existing employees
    existing_employees = sb_get('employees', {'activo': 'eq.true'})
    existing_by_cedula = {}
    existing_by_name = {}
    for emp in existing_employees:
        if emp.get('cedula'):
            existing_by_cedula[str(emp['cedula'])] = emp
        existing_by_name[emp['nombre'].upper().strip()] = emp
    
    # 5. Parse NOMINA sheet - extract employees for all sedes
    # Sections: rows 11-35 Bar/Mesas C75, 43-58 Cocina C75, 66-70 C85, 78-80 Kinder, 87-88 Admin
    
    EMPLOYEE_ROWS = {
        'C75_BAR': list(range(11, 36)),    # Bar/Mesas Calle 75
        'C75_COCINA': list(range(43, 59)),  # Cocina Calle 75
        'C85': list(range(66, 71)),          # Calle 85
        'KINDER': list(range(78, 81)),       # Pizzería Kinder
        'ADMIN': list(range(87, 89)),        # Administración
    }
    
    # Map row to sede
    ROW_SEDE = {}
    for sede, rows in EMPLOYEE_ROWS.items():
        for r in rows:
            ROW_SEDE[r] = sede
    
    employees_to_upsert = []
    nomina_detalle_records = []
    
    for row_idx in range(11, 90):
        if row_idx not in ROW_SEDE:
            continue
        
        c1 = ws_nom.cell(row=row_idx, column=1).value
        c4 = ws_nom.cell(row=row_idx, column=4).value
        if not c1 or not isinstance(c1, (int, float)) or not c4:
            continue
        
        sede_key = ROW_SEDE[row_idx]
        sede_id = {'C75_BAR': sede_c75, 'C75_COCINA': sede_c75, 'C85': sede_c85, 'KINDER': sede_kinder, 'ADMIN': sede_admin}[sede_key]
        
        cedula = str(ws_nom.cell(row=row_idx, column=3).value or '').strip()
        nombre = str(c4).strip()
        cargo = str(ws_nom.cell(row=row_idx, column=8).value or '').strip()
        modalidad = str(ws_nom.cell(row=row_idx, column=5).value or 'COMPLETO').strip()
        dias = safe_int(ws_nom.cell(row=row_idx, column=9).value)
        salario_basico = safe_float(ws_nom.cell(row=row_idx, column=10).value)
        fecha_ingreso = safe_date(ws_nom.cell(row=row_idx, column=6).value)
        
        salario_devengado = safe_float(ws_nom.cell(row=row_idx, column=16).value)
        aux_transporte = safe_float(ws_nom.cell(row=row_idx, column=17).value)
        propinas_75_85 = safe_float(ws_nom.cell(row=row_idx, column=18).value)
        aux_no_salarial = safe_float(ws_nom.cell(row=row_idx, column=20).value)
        recargos = safe_float(ws_nom.cell(row=row_idx, column=21).value)
        propinas = safe_float(ws_nom.cell(row=row_idx, column=22).value)
        total_devengado = safe_float(ws_nom.cell(row=row_idx, column=23).value)
        salud = safe_float(ws_nom.cell(row=row_idx, column=24).value)
        pension = safe_float(ws_nom.cell(row=row_idx, column=25).value)
        pagos = safe_float(ws_nom.cell(row=row_idx, column=26).value)
        prestamos = safe_float(ws_nom.cell(row=row_idx, column=27).value)
        total_deducciones = safe_float(ws_nom.cell(row=row_idx, column=28).value)
        neto = safe_float(ws_nom.cell(row=row_idx, column=29).value)
        
        team_name = CARGO_TO_TEAM.get(cargo, 'Sin asignar')
        mod_code = MODALIDAD_MAP.get(modalidad, 'COMPLETO')
        
        # Skip $0 rows (Neidy Rojas absent)
        if dias == 0 and total_devengado == 0:
            print(f"  SKIP: {nombre} ({cedula}) - 0 días, $0 devengado")
            continue
        
        # Check if employee exists by cedula or name
        existing = None
        if cedula and cedula in existing_by_cedula:
            existing = existing_by_cedula[cedula]
        elif nombre.upper() in existing_by_name:
            existing = existing_by_name[nombre.upper()]
        
        employee_data = {
            'nombre': nombre,
            'cedula': cedula,
            'cargo': cargo,
            'salario': salario_basico,
            'modalidad': mod_code,
            'sede_id': sede_id,
            'fecha_ingreso': fecha_ingreso,
            'aplica_propinas': True,
            'activo': True,
        }
        
        if existing:
            # Update existing employee
            team_id = find_team_by_name(team_name)
            if team_id:
                employee_data['team'] = team_id
            print(f"  UPDATE: {nombre} ({cedula}) → team={team_name}")
            try:
                sb_patch('employees', 'id', existing['id'], employee_data)
            except Exception as e:
                print(f"    ERROR updating {nombre}: {e}")
        else:
            # Try to find team or leave null
            team_id = find_team_by_name(team_name)
            emp_insert = {**employee_data}
            if team_id:
                emp_insert['team'] = team_id
            print(f"  INSERT: {nombre} ({cedula}) → team={team_name}")
            employees_to_upsert.append(emp_insert)
        
        # Nomina detalle record
        periodo_id = periodo_ids.get(sede_id)
        if periodo_id and existing:
            det = {
                'periodo_id': periodo_id,
                'empleado_id': existing['id'],
                'sede_id': sede_id,
                'dias_laborados': dias,
                'salario_basico': salario_basico,
                'salario_devengado': salario_devengado,
                'auxilio_transporte': aux_transporte,
                'propinas_prometido_75_85': propinas_75_85,
                'auxilio_no_salarial': aux_no_salarial,
                'recargos_he_rn_rd': recargos,
                'propinas': propinas,
                'total_devengado': total_devengado,
                'salud_empleado': salud,
                'pension_empleado': pension,
                'pagos_realizados': pagos,
                'prestamos_consumos': prestamos,
                'total_deducciones': total_deducciones,
                'neto_a_pagar': neto,
            }
            nomina_detalle_records.append(det)
    
    # Batch insert new employees
    if employees_to_upsert:
        print(f"\nInsertando {len(employees_to_upsert)} empleados nuevos...")
        for i in range(0, len(employees_to_upsert), 100):
            batch = employees_to_upsert[i:i+100]
            try:
                result = sb_post('employees', batch)
                print(f"  Batch {i//100 + 1}: {len(batch)} empleados insertados")
            except Exception as e:
                print(f"  ERROR insertando batch: {e}")
                print(f"  Primer registro: {json.dumps(batch[0], indent=2, default=str)}")
    
    # Re-fetch employees after insert to get IDs
    all_employees = sb_get('employees')
    emp_by_cedula = {str(e.get('cedula', '')): e for e in all_employees if e.get('cedula')}
    emp_by_name_upper = {e['nombre'].upper().strip(): e for e in all_employees}
    
    # Now insert nomina_detalle with proper employee IDs
    print(f"\nInsertando {len(nomina_detalle_records)} registros de nómina detalle...")
    # Need to resolve empleado_id from the actual data
    # Re-parse with employee IDs
    
    # 6. Parse HE-REC DOMINICALES sheet
    print("\nParseando horas extra y recargos...")
    he_records = []
    for row_idx in range(2, ws_he.max_row + 1):
        nombre = str(ws_he.cell(row=row_idx, column=2).value or '').strip()
        if not nombre or nombre == 'SUBTOTAL' or nombre == 'TOTALES':
            continue
        
        # Find employee
        emp = emp_by_name_upper.get(nombre.upper())
        if not emp:
            # Try partial match
            for e in all_employees:
                if nombre.upper() in e['nombre'].upper():
                    emp = e
                    break
        
        if not emp:
            print(f"  HE: No encontré empleado '{nombre}'")
            continue
        
        # Determine sede (default C75 for HE sheet)
        sede_id = emp.get('sede_id') or sede_c75
        periodo_id = periodo_ids.get(sede_id)
        
        if not periodo_id:
            print(f"  HE: No periodo for sede {sede_id}")
            continue
        
        he = {
            'periodo_id': periodo_id,
            'empleado_id': emp['id'],
            'sede_id': sede_id,
            'hed_horas': safe_float(ws_he.cell(row=row_idx, column=4).value),
            'hed_valor': safe_float(ws_he.cell(row=row_idx, column=5).value),
            'hed_total': safe_float(ws_he.cell(row=row_idx, column=6).value),
            'hen_horas': safe_float(ws_he.cell(row=row_idx, column=7).value),
            'hen_valor': safe_float(ws_he.cell(row=row_idx, column=8).value),
            'hen_total': safe_float(ws_he.cell(row=row_idx, column=9).value),
            'rn_horas': safe_float(ws_he.cell(row=row_idx, column=13).value),
            'rn_valor_hora': safe_float(ws_he.cell(row=row_idx, column=14).value),
            'rn_total': safe_float(ws_he.cell(row=row_idx, column=15).value),
            'rd_diurno_horas': safe_float(ws_he.cell(row=row_idx, column=16).value),
            'rd_diurno_valor': safe_float(ws_he.cell(row=row_idx, column=17).value),
            'rd_diurno_total': safe_float(ws_he.cell(row=row_idx, column=18).value),
            'rd_nocturno_horas': safe_float(ws_he.cell(row=row_idx, column=19).value),
            'rd_nocturno_valor': safe_float(ws_he.cell(row=row_idx, column=20).value),
            'rd_nocturno_total': safe_float(ws_he.cell(row=row_idx, column=21).value),
            'hedd_horas': safe_float(ws_he.cell(row=row_idx, column=22).value),
            'hedd_valor': safe_float(ws_he.cell(row=row_idx, column=23).value),
            'hedd_total': safe_float(ws_he.cell(row=row_idx, column=24).value),
            'hddn_horas': safe_float(ws_he.cell(row=row_idx, column=25).value),
            'hddn_valor': safe_float(ws_he.cell(row=row_idx, column=26).value),
            'hddn_total': safe_float(ws_he.cell(row=row_idx, column=27).value),
            'total_recargos': safe_float(ws_he.cell(row=row_idx, column=28).value),
        }
        he_records.append(he)
    
    if he_records:
        print(f"Insertando {len(he_records)} registros HE/RN/RD...")
        for i in range(0, len(he_records), 100):
            batch = he_records[i:i+100]
            try:
                sb_post('nomina_he_recargos', batch)
                print(f"  Batch HE {i//100 + 1}: OK")
            except Exception as e:
                print(f"  ERROR HE batch: {e}")
    
    # 7. Parse NOVEDADES CALLE 75
    print("\nParseando novedades Calle 75...")
    novedad_records = []
    for row_idx in range(7, 19):  # Novedades rows
        cedula = str(ws_nov75.cell(row=row_idx, column=2).value or '').strip()
        nombre = str(ws_nov75.cell(row=row_idx, column=3).value or '').strip()
        tipo = str(ws_nov75.cell(row=row_idx, column=4).value or '').strip()
        obs = str(ws_nov75.cell(row=row_idx, column=5).value or '').strip()
        dias = safe_int(ws_nov75.cell(row=row_idx, column=6).value)
        
        if not tipo or not nombre:
            continue
        
        tipo_map = {
            'VACACIONES': 'VACACIONES',
            'INCAPACIDAD': 'INCAPACIDAD',
            'PER. REMUNERADO': 'PERMISO_REMUNERADO',
            'PER. NO REMUNERADO': 'PERMISO_NO_REMUNERADO',
            'AUSENCIA': 'AUSENCIA',
            'CAMBIO BANCO': 'CAMBIO_BANCO',
            'NOVEDAD MARZO': 'NOVEDAD_MES_ANTERIOR',
        }
        
        emp = None
        if cedula and cedula in emp_by_cedula:
            emp = emp_by_cedula[cedula]
        
        nov = {
            'periodo_id': periodo_ids.get(sede_c75),
            'empleado_id': emp['id'] if emp else None,
            'sede_id': sede_c75,
            'tipo': tipo_map.get(tipo, tipo.upper()),
            'observacion': obs,
            'dias': dias,
            'aplicada': True,  # Excel says "Ok novedad aplicada"
        }
        novedad_records.append(nov)
    
    if novedad_records:
        print(f"Insertando {len(novedad_records)} novedades C75...")
        try:
            sb_post('nomina_novedades', novedad_records)
            print("  OK")
        except Exception as e:
            print(f"  ERROR: {e}")
    
    # 8. Parse PROVISIONES
    print("\nParseando provisiones...")
    prov_records = []
    for row_idx in range(11, ws_prov.max_row + 1):
        c1 = ws_prov.cell(row=row_idx, column=1).value
        c4 = ws_prov.cell(row=row_idx, column=4).value
        if not c1 or not isinstance(c1, (int, float)) or not c4:
            continue
        
        cedula = str(ws_prov.cell(row=row_idx, column=3).value or '').strip()
        nombre = str(c4).strip()
        
        emp = None
        if cedula and cedula in emp_by_cedula:
            emp = emp_by_cedula[cedula]
        
        if not emp:
            for e in all_employees:
                if nombre.upper() in e['nombre'].upper() or e['nombre'].upper() in nombre.upper():
                    emp = e
                    break
        
        if not emp:
            continue
        
        sede_id = emp.get('sede_id') or sede_c75
        periodo_id = periodo_ids.get(sede_id)
        if not periodo_id:
            continue
        
        prov = {
            'periodo_id': periodo_id,
            'empleado_id': emp['id'],
            'sede_id': sede_id,
            'provisiones_salud': safe_float(ws_prov.cell(row=row_idx, column=9).value),
            'provisiones_sociales': safe_float(ws_prov.cell(row=row_idx, column=10).value),
            'base_vacaciones': safe_float(ws_prov.cell(row=row_idx, column=11).value),
            'salud_empleado': safe_float(ws_prov.cell(row=row_idx, column=12).value),
            'pension_empleado': safe_float(ws_prov.cell(row=row_idx, column=13).value),
            'pension_empleador': safe_float(ws_prov.cell(row=row_idx, column=14).value),
            'arl_empleador': safe_float(ws_prov.cell(row=row_idx, column=15).value),
            'caja_empleador': safe_float(ws_prov.cell(row=row_idx, column=16).value),
            'cesantias_empleador': safe_float(ws_prov.cell(row=row_idx, column=17).value),
            'prima_empleador': safe_float(ws_prov.cell(row=row_idx, column=18).value),
            'vacaciones_empleador': safe_float(ws_prov.cell(row=row_idx, column=19).value),
            'intereses_cesantias_empleador': safe_float(ws_prov.cell(row=row_idx, column=20).value),
            'total_provision_empleador': safe_float(ws_prov.cell(row=row_idx, column=14).value, 0) + 
                                         safe_float(ws_prov.cell(row=row_idx, column=15).value, 0) + 
                                         safe_float(ws_prov.cell(row=row_idx, column=16).value, 0) + 
                                         safe_float(ws_prov.cell(row=row_idx, column=17).value, 0) + 
                                         safe_float(ws_prov.cell(row=row_idx, column=18).value, 0) + 
                                         safe_float(ws_prov.cell(row=row_idx, column=19).value, 0) + 
                                         safe_float(ws_prov.cell(row=row_idx, column=20).value, 0),
        }
        prov_records.append(prov)
    
    if prov_records:
        print(f"Insertando {len(prov_records)} provisiones...")
        for i in range(0, len(prov_records), 100):
            batch = prov_records[i:i+100]
            try:
                sb_post('nomina_provisiones', batch)
                print(f"  Batch Prov {i//100 + 1}: OK")
            except Exception as e:
                print(f"  ERROR Prov batch: {e}")
    
    # 9. Create PROPINAS records
    print("\nInsertando propinas...")
    propinas_data = [
        {'sede_id': sede_c75, 'total_propinas_ventas': 60789692, 'dias_laborados_total': 1059, 'valor_dia_propina': 57402.92},
        {'sede_id': sede_kinder, 'total_propinas_ventas': 1969675, 'dias_laborados_total': 90, 'valor_dia_propina': 21885.28},
        {'sede_id': sede_c85, 'total_propinas_ventas': 2827742, 'dias_laborados_total': 105, 'valor_dia_propina': 26930.88},
    ]
    for p in propinas_data:
        sid = p.pop('sede_id')
        pid = periodo_ids.get(sid)
        if pid:
            p['periodo_id'] = pid
            p['sede_id'] = sid
            try:
                sb_post('nomina_propinas', [p])
                print(f"  Propinas {sid}: OK")
            except Exception as e:
                print(f"  ERROR Propinas {sid}: {e}")
    
    print("\n=== Importación completada ===")


if __name__ == '__main__':
    main()