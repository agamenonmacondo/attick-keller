#!/usr/bin/env python3
"""Import nómina abril 2026 — A&K Supabase (Calle 75)"""
import psycopg2
import openpyxl
from decimal import Decimal, ROUND_HALF_UP
import time

DB_PASS = 'Pita5721@153'
XLSX = '/home/mod/.hermes/cache/documents/doc_8319112453a1_Nomina atico mes de abril  2026.xlsx'

def dec(val):
    if val is None: return Decimal('0')
    try: return Decimal(str(val).replace(',','.')).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    except: return Decimal('0')

def ced(val):
    if val is None: return None
    try: return str(int(float(str(val))))
    except: return str(val).strip()

def connect(retries=5):
    for i in range(retries):
        try:
            conn = psycopg2.connect(
                host='aws-1-us-east-2.pooler.supabase.com', port=5432,
                database='postgres', user='postgres.pbllaipsdfypelnwrvpy',
                password=DB_PASS, sslmode='require', connect_timeout=15)
            conn.autocommit = True
            return conn
        except Exception as e:
            print(f"  Retry {i+1}/{retries}: {str(e)[:60]}")
            time.sleep(3)
    raise RuntimeError("Could not connect after retries")

# Connect
conn = connect()
cur = conn.cursor()
print("Connected!")

# Periodo IDs
cur.execute("SELECT id, sede FROM public.nomina_periodos WHERE periodo='ABRIL 2026';")
periodos = {r[1]: str(r[0]) for r in cur.fetchall()}
print(f"Periodos: {periodos}")

# Staff
cur.execute("SELECT id, cedula, nombre_completo FROM public.pos_nomina_staff;")
staff_by_ced = {}
for r in cur.fetchall():
    c = ced(r[1])
    if c: staff_by_ced[c] = {'id': str(r[0]), 'nombre': r[2]}
print(f"Staff: {len(staff_by_ced)} by cedula")

wb = openpyxl.load_workbook(XLSX, data_only=True)

# ============================================================
# 1. NOMINA DETALLE C75
# ============================================================
ws = wb['NOMINA MES DE ABRIL 2026']
c75_id = periodos['C75']
staff_upd = det_ins = 0
not_found = []

for row in range(9, 58):
    c1 = ws.cell(row=row, column=1).value
    c3 = ws.cell(row=row, column=3).value
    c4 = ws.cell(row=row, column=4).value
    c4s = str(c4 or '').strip().upper()
    if not c3 and not c4: continue
    if 'TOTAL' in c4s or 'SUBTOTAL' in c4s: continue
    if str(c3 or '').strip().upper() == 'CEDULA': continue
    
    cedula = ced(c3)
    if not cedula: continue
    s = staff_by_ced.get(cedula)
    if not s:
        not_found.append((cedula, str(c4 or '')[:25]))
        continue

    sid = s['id']
    nombre = str(c4 or '').strip()
    modalidad = str(ws.cell(row=row, column=5).value or '').strip() or 'COMPLETO'
    cargo = str(ws.cell(row=row, column=8).value or '').strip() or None
    dias = int(float(str(ws.cell(row=row, column=9).value or '30')))
    sal_bas = dec(ws.cell(row=row, column=10).value)
    sal_dev = dec(ws.cell(row=row, column=16).value)
    aux_tra = dec(ws.cell(row=row, column=17).value)
    pp85 = dec(ws.cell(row=row, column=18).value)
    pp75 = dec(ws.cell(row=row, column=19).value)
    aux_ns = dec(ws.cell(row=row, column=20).value)
    recargos = dec(ws.cell(row=row, column=21).value)
    propinas = dec(ws.cell(row=row, column=22).value)
    tot_dev = dec(ws.cell(row=row, column=23).value)
    salud = dec(ws.cell(row=row, column=24).value)
    pension = dec(ws.cell(row=row, column=25).value)
    pagos = dec(ws.cell(row=row, column=26).value)
    prestamos = dec(ws.cell(row=row, column=27).value)
    tot_ded = dec(ws.cell(row=row, column=28).value)
    neto = dec(ws.cell(row=row, column=29).value)
    ibc = ws.cell(row=row, column=30).value
    
    cur.execute("""UPDATE public.pos_nomina_staff 
        SET salario=%s, cargo=%s, modalidad=%s, sede=%s, aplica_propinas=true, auxilio_no_salarial=%s
        WHERE id=%s""", (sal_bas, cargo, modalidad, 'C75', aux_ns, sid))
    staff_upd += 1
    
    cur.execute("""INSERT INTO public.nomina_detalle (
        periodo_id,staff_id,sede,dias_laborados,salario_basico,salario_devengado,
        auxilio_transporte,propinas_prometido_75_85,propinas_prometido_75,
        auxilio_no_salarial,recargos_he_rn_rd,propinas,total_devengado,
        salud_empleado,pension_empleado,pagos_realizados,prestamos_consumos,
        total_deducciones,neto_a_pagar,ibc
    ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
    ON CONFLICT (periodo_id,staff_id,sede) DO UPDATE SET
        dias_laborados=EXCLUDED.dias_laborados,total_devengado=EXCLUDED.total_devengado,
        total_deducciones=EXCLUDED.total_deducciones,neto_a_pagar=EXCLUDED.neto_a_pagar""",
        (c75_id,sid,'C75',dias,sal_bas,sal_dev,aux_tra,pp85,pp75,aux_ns,recargos,propinas,tot_dev,
         salud,pension,pagos,prestamos,tot_ded,neto,dec(ibc) if ibc else None))
    det_ins += 1

# Fecha ingreso
for row in range(9, 58):
    c3 = ws.cell(row=row, column=3).value
    c6 = ws.cell(row=row, column=6).value
    if not c3 or not c6: continue
    cedula = ced(c3)
    s = staff_by_ced.get(cedula)
    if not s: continue
    fecha = c6
    if isinstance(fecha, str):
        from datetime import datetime
        for fmt in ['%d/%m/%Y','%Y-%m-%d','%m/%d/%Y']:
            try: fecha = datetime.strptime(fecha, fmt).date(); break
            except: pass
    elif hasattr(fecha, 'date'):
        fecha = fecha.date() if callable(getattr(fecha,'date',None)) else fecha
    try:
        cur.execute("UPDATE public.pos_nomina_staff SET fecha_ingreso=%s WHERE id=%s", (fecha, s['id']))
    except: pass

# Update periodo totals
cur.execute("""UPDATE public.nomina_periodos SET
    total_devengado=(SELECT COALESCE(SUM(total_devengado),0) FROM nomina_detalle WHERE periodo_id=%s AND sede='C75'),
    total_deducciones=(SELECT COALESCE(SUM(total_deducciones),0) FROM nomina_detalle WHERE periodo_id=%s AND sede='C75'),
    total_neto=(SELECT COALESCE(SUM(neto_a_pagar),0) FROM nomina_detalle WHERE periodo_id=%s AND sede='C75')
    WHERE id=%s""", (c75_id,c75_id,c75_id,c75_id))

print(f"1. C75 Detalle: {det_ins} rows, {staff_upd} staff updated")
if not_found: print(f"   Not found ({len(not_found)}): {not_found[:5]}")

# ============================================================
# 2. HE/RECARGOS
# ============================================================
ws_he = wb['HE - REC DOMINICALES ABRIL 2026']
he_ins = 0

# Sections: C75 rows 2-36, C85 rows 41-47, KINDER rows 52-56
sections = [('C75', 2, 36), ('C85', 41, 47), ('KINDER', 52, 56)]

for sede, start, end in sections:
    pid = periodos[sede]
    for row in range(start, end+1):
        col1 = ws_he.cell(row=row, column=1).value
        col2 = ws_he.cell(row=row, column=2).value
        if not col1 or not col2: continue
        try: int(float(str(col1)))
        except: continue
        
        nombre = str(col2).strip().upper()
        # Find staff by name fuzzy
        sid = None
        for c, s in staff_by_ced.items():
            if s['nombre'].upper() in nombre or nombre in s['nombre'].upper():
                sid = s['id']
                break
        if not sid: continue
        
        cur.execute("""INSERT INTO public.nomina_he_recargos (
            periodo_id,staff_id,sede,
            hed_horas,hed_valor,hed_total,
            hen_horas,hen_valor,hen_total,
            rn_horas,rn_valor_hora,rn_total,
            rd_diurno_horas,rd_diurno_valor,rd_diurno_total,
            rd_nocturno_horas,rd_nocturno_valor,rd_nocturno_total,
            hedd_horas,hedd_valor,hedd_total,
            hddn_horas,hddn_valor,hddn_total,
            total_recargos
        ) VALUES (%s,%s,%s, %s,%s,%s, %s,%s,%s, %s,%s,%s, %s,%s,%s, %s,%s,%s, %s,%s,%s, %s,%s,%s, %s)
        ON CONFLICT (periodo_id,staff_id,sede) DO UPDATE SET
            total_recargos=EXCLUDED.total_recargos""",
        (pid,sid,sede,
         dec(ws_he.cell(row=row,column=4).value),dec(ws_he.cell(row=row,column=5).value),dec(ws_he.cell(row=row,column=6).value),
         dec(ws_he.cell(row=row,column=7).value),dec(ws_he.cell(row=row,column=8).value),dec(ws_he.cell(row=row,column=9).value),
         dec(ws_he.cell(row=row,column=10).value),dec(ws_he.cell(row=row,column=11).value),dec(ws_he.cell(row=row,column=12).value),
         dec(ws_he.cell(row=row,column=13).value),dec(ws_he.cell(row=row,column=14).value),dec(ws_he.cell(row=row,column=15).value),
         dec(ws_he.cell(row=row,column=16).value),dec(ws_he.cell(row=row,column=17).value),dec(ws_he.cell(row=row,column=18).value),
         dec(ws_he.cell(row=row,column=19).value),dec(ws_he.cell(row=row,column=20).value),dec(ws_he.cell(row=row,column=21).value),
         dec(ws_he.cell(row=row,column=22).value),dec(ws_he.cell(row=row,column=23).value),dec(ws_he.cell(row=row,column=24).value),
         dec(ws_he.cell(row=row,column=25).value),dec(ws_he.cell(row=row,column=26).value),dec(ws_he.cell(row=row,column=27).value),
         dec(ws_he.cell(row=row,column=28).value)))
        he_ins += 1

print(f"2. HE/Recargos: {he_ins} rows")

# ============================================================
# 3. NOVEDADES C75
# ============================================================
ws_nov = wb['NOVEDADES CALLE 75 ABRIL 26']
nov_ins = 0

# Scan for data rows (skip headers)
for row in range(1, ws_nov.max_row + 1):
    c1 = ws_nov.cell(row=row, column=1).value
    c4 = ws_nov.cell(row=row, column=4).value
    c5 = ws_nov.cell(row=row, column=5).value
    c6 = ws_nov.cell(row=row, column=6).value
    c7 = ws_nov.cell(row=row, column=7).value
    
    c1s = str(c1 or '').strip()
    if c1s.upper() in ('N°','NO','N','INGRESO','NOVEDADES','') or c1s == 'None': continue
    
    # Check if it's a novedad row or an ingreso row
    c2 = ws_nov.cell(row=row, column=2).value  # CEDULA
    c3 = ws_nov.cell(row=row, column=3).value  # NOMBRE
    
    cedula = ced(c2)
    if not cedula: continue
    s = staff_by_ced.get(cedula)
    if not s: continue
    
    tipo = str(c4 or '').strip() if c4 else 'INGRESO'
    fecha_ing = c5
    banco = str(c6 or '').strip() if c6 else None
    cuenta = str(c7 or '').strip() if c7 else None
    
    # Update bank info if present
    if banco or cuenta:
        try:
            cur.execute("UPDATE pos_nomina_staff SET banco=%s, cuenta_bancaria=%s WHERE id=%s",
                       (banco, cuenta, s['id']))
        except: pass

print(f"3. Novedades processed")

# ============================================================
# 4. PROVISIONES
# ============================================================
ws_prov = wb['PROVISIONES ABRIL 2026']
prov_ins = 0

# Read structure - prov has sections by sede
# Find header row
header_row = None
for r in range(1, 10):
    val = str(ws_prov.cell(row=r, column=1).value or '').strip().lower()
    if 'cedula' in val or 'nombre' in val:
        header_row = r
        break

# Scan for C75 section
# Let me print structure first
print(f"\nProvisiones sheet: {ws_prov.max_row} rows x {ws_prov.max_column} cols")
print("First 10 rows:")
for r in range(1, min(11, ws_prov.max_row + 1)):
    row_data = {}
    for c in range(1, min(28, ws_prov.max_column + 1)):
        val = ws_prov.cell(row=r, column=c).value
        if val is not None:
            row_data[c] = str(val)[:40]
    if row_data:
        print(f"  Row {r}: {row_data}")

cur.close()
conn.close()
print("\nPREVIEW DONE — ready for full provisiones import")