#!/usr/bin/env python3
"""
Importar nómina abril 2026 a A&K vía Next.js API route
No requiere SUPABASE_SERVICE_ROLE_KEY directamente — usa la app como proxy.

PRERREQUISITOS:
  1. Ejecutar nomina_ddl_ak.sql en Supabase Dashboard SQL Editor
  2. Tener dev server corriendo: cd /mnt/f/attick-keller/web && npm run dev
  3. pip install openpyxl requests

Uso:
  python3 scripts/import_nomina_ak.py [--port 3000] [--base-url http://localhost:3000]
"""

import openpyxl
import requests
import sys
import argparse
from datetime import datetime
from typing import Optional

# ── Args ────────────────────────────────────────────
parser = argparse.ArgumentParser()
parser.add_argument('--port', type=int, default=3000)
parser.add_argument('--base-url', type=str, default=None)
args = parser.parse_args()

BASE_URL = args.base_url or f'http://localhost:{args.port}'
API_URL = f'{BASE_URL}/api/admin/nomina-import'
EXCEL_PATH = '/home/mod/.hermes/cache/documents/doc_8319112453a1_Nomina atico mes de abril  2026.xlsx'

# ── Auth cookie (from browser session) ──────────────
# For dev, we'll use the anon approach — the API route checks admin auth
# In production, you'd pass a session cookie. For local dev, we need to 
# configure the API route to accept a secret token or bypass auth.
SECRET_TOKEN='nomina...2026'  # Must match route.ts

HEADERS = {
    'Content-Type': 'application/json',
    'X-Import-Token': SECRET_TOKEN,
}

# ── Helpers ──────────────────────────────────────────
def safe_float(val, default=0):
    if val is None: return default
    try: return round(float(val), 2)
    except (ValueError, TypeError): return default

def safe_int(val, default=0):
    if val is None: return default
    try: return int(float(val))
    except (ValueError, TypeError): return default

def api_call(action: str, data: dict | list) -> dict:
    """Call the Next.js API route"""
    payload = {'action': action, 'data': data}
    r = requests.post(API_URL, json=payload, headers=HEADERS, timeout=30)
    if r.status_code >= 400:
        print(f"  ✗ {action} → {r.status_code}: {r.text[:300]}")
        return {'error': r.text}
    return r.json()

# ═══════════════════════════════════════════════════════
# 1. PARSE EXCEL
# ═══════════════════════════════════════════════════════
print("📖 Leyendo archivo Excel...")
wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
ws1 = wb['NOMINA MES DE ABRIL 2026']

sections = {
    'BAR_Y_MESAS_C75': (7, 35),
    'COCINA_C75': (39, 58),
    'CALLE_85': (62, 70),
    'KINDER': (74, 80),
    'ADMIN': (83, 88),
}

def parse_nomina_section(ws, start_row, end_row):
    employees = []
    for row_num in range(start_row, end_row + 1):
        cedula = ws.cell(row=row_num, column=3).value
        nombre = ws.cell(row=row_num, column=4).value
        
        if cedula is None and nombre is None: continue
        if isinstance(cedula, str) and cedula.strip().upper() == 'CEDULA': continue
        if isinstance(nombre, str) and nombre.strip().upper() in ('NOMBRE DEL TRABAJADOR', 'TOTAL DEVENGADO', 'TOTALES DE NOMINAS'): continue
        
        empleado = {
            'cedula': str(int(cedula)) if isinstance(cedula, (int, float)) else str(cedula),
            'tipo_doc': str(ws.cell(row=row_num, column=2).value or 'CC').strip(),
            'nombre': str(nombre or '').strip(),
            'modalidad': str(ws.cell(row=row_num, column=5).value or 'COMPLETO').strip(),
            'fecha_ingreso': ws.cell(row=row_num, column=6).value,
            'fecha_retiro': ws.cell(row=row_num, column=7).value,
            'cargo': str(ws.cell(row=row_num, column=8).value or '').strip(),
            'dias': safe_int(ws.cell(row=row_num, column=9).value),
            'salario_basico': safe_float(ws.cell(row=row_num, column=10).value),
            'pagos_acordados': safe_float(ws.cell(row=row_num, column=11).value),
            'aplica_propinas': str(ws.cell(row=row_num, column=12).value or 'SI').strip().upper() == 'SI',
            'aplica_propinas_dias': safe_int(ws.cell(row=row_num, column=13).value),
            'prop_c75_dias': safe_int(ws.cell(row=row_num, column=14).value, None),
            'prop_c85_dias': safe_int(ws.cell(row=row_num, column=15).value, None),
            'salario': safe_float(ws.cell(row=row_num, column=16).value),
            'aux_transporte': safe_float(ws.cell(row=row_num, column=17).value),
            'propinas_prom_75_85': safe_float(ws.cell(row=row_num, column=18).value),
            'propinas_prom_75': safe_float(ws.cell(row=row_num, column=19).value),
            'aux_no_salarial': safe_float(ws.cell(row=row_num, column=20).value),
            'recargos_he_rn_rd': safe_float(ws.cell(row=row_num, column=21).value),
            'propinas': safe_float(ws.cell(row=row_num, column=22).value),
            'total_devengado': safe_float(ws.cell(row=row_num, column=23).value),
            'salud': safe_float(ws.cell(row=row_num, column=24).value),
            'pension': safe_float(ws.cell(row=row_num, column=25).value),
            'pagos_realizados': safe_float(ws.cell(row=row_num, column=26).value),
            'prestamos_consumos': safe_float(ws.cell(row=row_num, column=27).value),
            'total_deducciones': safe_float(ws.cell(row=row_num, column=28).value),
            'neto_a_pagar': safe_float(ws.cell(row=row_num, column=29).value),
        }
        
        if isinstance(empleado['fecha_ingreso'], datetime):
            empleado['fecha_ingreso'] = empleado['fecha_ingreso'].strftime('%Y-%m-%d')
        if isinstance(empleado['fecha_retiro'], datetime):
            empleado['fecha_retiro'] = empleado['fecha_retiro'].strftime('%Y-%m-%d')
        
        if empleado['nombre'] and empleado['cedula']:
            employees.append(empleado)
    
    return employees

all_employees = {}
for section_name, (start, end) in sections.items():
    sede = 'C75' if 'C75' in section_name else ('C85' if 'C85' in section_name else ('KINDER' if 'KINDER' in section_name else 'ADMIN'))
    employees = parse_nomina_section(ws1, start, end)
    for emp in employees:
        emp['sede'] = sede
        key = f"{emp['cedula']}_{sede}"
        if key not in all_employees:
            all_employees[key] = emp
        else:
            existing = all_employees[key]
            for field in ['dias', 'salario', 'total_devengado', 'total_deducciones', 'neto_a_pagar',
                         'salud', 'pension', 'propinas', 'recargos_he_rn_rd', 'pagos_realizados',
                         'prestamos_consumos', 'aux_transporte', 'propinas_prom_75_85', 'propinas_prom_75',
                         'aux_no_salarial']:
                existing[field] += emp[field]

print(f"  Empleados únicos: {len(all_employees)}")

# ═══════════════════════════════════════════════════════
# 2. PARSE HE/RECARGOS
# ═══════════════════════════════════════════════════════
ws2 = wb['HE - REC DOMINICALES ABRIL 2026']
he_records = []
current_sede = 'C75'

for row_num in range(2, ws2.max_row + 1):
    nombre = ws2.cell(row=row_num, column=2).value
    if nombre and isinstance(nombre, str):
        upper = nombre.upper().strip()
        if 'CALLE 85' in upper: current_sede = 'C85'; continue
        if 'KINDER' in upper: current_sede = 'KINDER'; continue
        if 'ADMIN' in upper: current_sede = 'ADMIN'; continue
        if 'TOTAL' in upper or 'SUBTOTAL' in upper: continue
        if 'EMPLEADO' in upper: continue
    
    if not nombre or not str(nombre).strip(): continue
    num = ws2.cell(row=row_num, column=1).value
    if num is None and ws2.cell(row=row_num, column=3).value is None: continue
    
    he_records.append({
        'nombre': str(nombre).strip(),
        'sede': current_sede,
        'hed_horas': safe_float(ws2.cell(row=row_num, column=4).value),
        'hed_valor': safe_float(ws2.cell(row=row_num, column=5).value),
        'hed_total': safe_float(ws2.cell(row=row_num, column=6).value),
        'hen_horas': safe_float(ws2.cell(row=row_num, column=7).value),
        'hen_valor': safe_float(ws2.cell(row=row_num, column=8).value),
        'hen_total': safe_float(ws2.cell(row=row_num, column=9).value),
        'rn_horas': safe_float(ws2.cell(row=row_num, column=13).value),
        'rn_valor_hora': safe_float(ws2.cell(row=row_num, column=14).value),
        'rn_total': safe_float(ws2.cell(row=row_num, column=15).value),
        'rd_diurno_horas': safe_float(ws2.cell(row=row_num, column=16).value),
        'rd_diurno_valor': safe_float(ws2.cell(row=row_num, column=17).value),
        'rd_diurno_total': safe_float(ws2.cell(row=row_num, column=18).value),
        'rd_nocturno_horas': safe_float(ws2.cell(row=row_num, column=19).value),
        'rd_nocturno_valor': safe_float(ws2.cell(row=row_num, column=20).value),
        'rd_nocturno_total': safe_float(ws2.cell(row=row_num, column=21).value),
        'hedd_horas': safe_float(ws2.cell(row=row_num, column=22).value),
        'hedd_valor': safe_float(ws2.cell(row=row_num, column=23).value),
        'hedd_total': safe_float(ws2.cell(row=row_num, column=24).value),
        'hddn_horas': safe_float(ws2.cell(row=row_num, column=25).value),
        'hddn_valor': safe_float(ws2.cell(row=row_num, column=26).value),
        'hddn_total': safe_float(ws2.cell(row=row_num, column=27).value),
        'total_recargos': safe_float(ws2.cell(row=row_num, column=28).value),
    })

print(f"  HE/Recargos: {len(he_records)}")

# ═══════════════════════════════════════════════════════
# 3. PARSE NOVEDADES
# ═══════════════════════════════════════════════════════
novedades = []
ws_nov75 = wb['NOVEDADES CALLE 75 ABRIL 26']
in_novedades = False
in_ingresos = False

for row_num in range(1, ws_nov75.max_row + 1):
    col1 = ws_nov75.cell(row=row_num, column=1).value
    if isinstance(col1, str) and 'INGRESO' in col1.upper(): in_ingresos = True; in_novedades = False; continue
    if isinstance(col1, str) and 'NOVEDADES' in col1.upper(): in_novedades = True; in_ingresos = False; continue
    if isinstance(col1, str) and 'RETIRO' in col1.upper(): in_novedades = False; in_ingresos = False; continue
    if isinstance(col1, str) and col1.strip().upper() == 'N°': continue
    
    if (in_novedades or in_ingresos) and col1 and isinstance(col1, (int, float)):
        cedula = ws_nov75.cell(row=row_num, column=2).value
        nombre = ws_nov75.cell(row=row_num, column=3).value or ''
        tipo = 'INGRESO' if in_ingresos else str(ws_nov75.cell(row=row_num, column=4).value or '').strip().upper()
        obs = str(ws_nov75.cell(row=row_num, column=5).value or '').strip() if in_novedades else str(ws_nov75.cell(row=row_num, column=4).value or '')
        dias = ws_nov75.cell(row=row_num, column=6 + (0 if in_novedades else 1)).value
        
        novedades.append({
            'cedula': str(int(cedula)) if isinstance(cedula, (int, float)) else str(cedula),
            'nombre': str(nombre).strip(),
            'tipo': tipo,
            'observacion': obs,
            'dias': str(dias) if dias else None,
            'sede': 'C75',
        })

print(f"  Novedades: {len(novedades)}")

# ═══════════════════════════════════════════════════════
# 4. PARSE PROVISIONES
# ═══════════════════════════════════════════════════════
ws_prov = wb['PROVISIONES ABRIL 2026']
provisiones = []
current_sede = 'C75'

for row_num in range(7, ws_prov.max_row + 1):
    col4 = ws_prov.cell(row=row_num, column=4).value
    if isinstance(col4, str):
        u = col4.strip().upper()
        if 'BAR Y MESAS' in u: current_sede = 'C75'; continue
        elif 'COCINA' in u and 'C75' not in u: current_sede = 'C75'; continue
        elif 'CALLE 85' in u or 'PERSONAL CALLE 85' in u: current_sede = 'C85'; continue
        elif 'KINDER' in u: current_sede = 'KINDER'; continue
        elif 'ADMINISTRATIVO' in u: current_sede = 'ADMIN'; continue
        elif 'TOTAL' in u: continue
    
    col3 = ws_prov.cell(row=row_num, column=3).value
    if col3 is None or (isinstance(col3, str) and col3.strip().upper() == 'CEDULA'): continue
    nombre = ws_prov.cell(row=row_num, column=4).value
    if not nombre or not isinstance(nombre, str) or 'TOTAL' in nombre.upper(): continue
    
    cedula = col3
    provisiones.append({
        'cedula': str(int(cedula)) if isinstance(cedula, (int, float)) else str(cedula),
        'nombre': str(nombre).strip(),
        'sede': current_sede,
        'provisiones_salud': safe_float(ws_prov.cell(row=row_num, column=9).value),
        'provisiones_sociales': safe_float(ws_prov.cell(row=row_num, column=10).value),
        'base_vacaciones': safe_float(ws_prov.cell(row=row_num, column=11).value),
        'salud_empleado': safe_float(ws_prov.cell(row=row_num, column=12).value),
        'pension_empleado': safe_float(ws_prov.cell(row=row_num, column=13).value),
        'pension_empleador': safe_float(ws_prov.cell(row=row_num, column=14).value),
        'arl_empleador': safe_float(ws_prov.cell(row=row_num, column=15).value),
        'caja_empleador': safe_float(ws_prov.cell(row=row_num, column=16).value),
        'cesantias_empleador': safe_float(ws_prov.cell(row=row_num, column=17).value),
        'prima_empleador': safe_float(ws_prov.cell(row=row_num, column=18).value),
        'vacaciones_empleador': safe_float(ws_prov.cell(row=row_num, column=19).value),
        'intereses_cesantias_empleador': safe_float(ws_prov.cell(row=row_num, column=20).value),
    })

print(f"  Provisiones: {len(provisiones)}")

# ═══════════════════════════════════════════════════════
# 5. CHECK STATUS
# ═══════════════════════════════════════════════════════
print(f"\n🔍 Verificando estado de tablas vía API...")
print(f"  URL: {API_URL}?action=status")

try:
    r = requests.get(f'{API_URL}?action=status', headers=HEADERS, timeout=10)
    if r.status_code >= 400:
        print(f"  ✗ Error {r.status_code}: {r.text[:300]}")
        print("\n⚠️  Asegurate de que:")
        print("  1. El DDL (nomina_ddl_ak.sql) fue ejecutado en Supabase Dashboard")
        print("  2. El dev server está corriendo (npm run dev)")
        print("  3. La API route es accesible")
        sys.exit(1)
    
    status = r.json()
    print("  Estado de tablas:")
    for table in ['pos_nomina_staff', 'nomina_periodos', 'nomina_detalle', 'nomina_he_recargos', 'nomina_novedades', 'nomina_provisiones', 'nomina_propinas']:
        info = status.get(table, {})
        if info.get('exists'):
            print(f"    ✓ {table}: existe ({info.get('count', '?')} rows)")
        else:
            print(f"    ✗ {table}: NO EXISTE — ejecutar DDL primero")
    
    staff_list = status.get('staff_list', [])
    staff_id_map = {s['cedula']: s['id'] for s in staff_list}
    print(f"\n  Staff en BD: {len(staff_list)} empleados")
    
    periodos = status.get('periodos', [])
    periodo_by_sede = {p['sede']: p['id'] for p in periodos}
    print(f"  Periodos: {periodos}")
except Exception as e:
    print(f"  ✗ No se pudo conectar a la API: {e}")
    print("\n⚠️  Asegurate de que el dev server esté corriendo (npm run dev)")
    sys.exit(1)

# ═══════════════════════════════════════════════════════
# 6. UPSERT EMPLOYEES
# ═══════════════════════════════════════════════════════
print("\n📝 Subiendo empleados...")
empleados_to_upsert = []
for key, emp in all_employees.items():
    empleados_to_upsert.append({
        'cedula': emp['cedula'],
        'nombre_completo': emp['nombre'],
        'salario': emp['salario_basico'],
        'cargo': emp['cargo'],
        'modalidad': emp['modalidad'],
        'sede': emp['sede'],
        'aplica_propinas': emp['aplica_propinas'],
        'fecha_ingreso': emp['fecha_ingreso'],
        'es_medio_tiempo': 'MEDIO' in emp['modalidad'].upper(),
        'auxilio_no_salarial': emp['aux_no_salarial'],
    })

# Send in batches of 20
batch_size = 20
for i in range(0, len(empleados_to_upsert), batch_size):
    batch = empleados_to_upsert[i:i+batch_size]
    result = api_call('upsert_staff', batch)
    if 'results' in result:
        for r in result['results']:
            status = r.get('status', 'unknown')
            if status == 'error':
                print(f"    ✗ {r['cedula']}: {r.get('error', 'unknown error')}")
            else:
                staff_id_map[r['cedula']] = r['id']
    print(f"  Batch {i//batch_size+1}/{(len(empleados_to_upsert)-1)//batch_size+1} procesado")

print(f"  Staff en mapa: {len(staff_id_map)} cédulas")

# ═══════════════════════════════════════════════════════
# 7. INSERT NOMINA_DETALLE
# ═══════════════════════════════════════════════════════
print("\n📊 Insertando nómina detalle...")
detalle_records = []
missing = []
for key, emp in all_employees.items():
    cedula = emp['cedula']
    sede = emp['sede']
    periodo_id = periodo_by_sede.get(sede)
    staff_id = staff_id_map.get(cedula)
    
    if not periodo_id:
        missing.append(f"periodo para sede {sede}")
        continue
    if not staff_id:
        missing.append(f"staff_id para {cedula} ({emp['nombre']})")
        continue
    
    detalle_records.append({
        'periodo_id': periodo_id,
        'staff_id': staff_id,
        'sede': sede,
        'dias_laborados': emp['dias'],
        'salario_basico': emp['salario_basico'],
        'salario_devengado': emp['salario'],
        'auxilio_transporte': emp['aux_transporte'],
        'propinas_prometido_75_85': emp['propinas_prom_75_85'],
        'propinas_prometido_75': emp['propinas_prom_75'],
        'auxilio_no_salarial': emp['aux_no_salarial'],
        'recargos_he_rn_rd': emp['recargos_he_rn_rd'],
        'propinas': emp['propinas'],
        'total_devengado': emp['total_devengado'],
        'salud_empleado': emp['salud'],
        'pension_empleado': emp['pension'],
        'pagos_realizados': emp['pagos_realizados'],
        'prestamos_consumos': emp['prestamos_consumos'],
        'total_deducciones': emp['total_deducciones'],
        'neto_a_pagar': emp['neto_a_pagar'],
    })

if missing:
    print(f"  ⚠️ Faltan: {missing[:5]}")
print(f"  Registros: {len(detalle_records)}")

# Insert in batches of 20
for i in range(0, len(detalle_records), batch_size):
    batch = detalle_records[i:i+batch_size]
    result = api_call('insert_detalle', batch)
    print(f"  Batch {i//batch_size+1}: {result}")

# ═══════════════════════════════════════════════════════
# 8. INSERT HE/RECARGOS
# ═══════════════════════════════════════════════════════
print("\n⏱️ Insertando HE/Recargos...")

# Build name-to-cedula map from all_employees
name_to_cedula = {}
for emp in all_employees.values():
    name_to_cedula[emp['nombre'].upper().strip()] = emp['cedula']
    parts = emp['nombre'].split()
    if len(parts) >= 2:
        name_to_cedula[parts[-1].upper()] = emp['cedula']

he_records_db = []
for rec in he_records:
    nombre_upper = rec['nombre'].upper().strip()
    cedula = None
    
    for nk, nc in name_to_cedula.items():
        if nombre_upper == nk or nombre_upper.startswith(nk[:20]) or nk.startswith(nombre_upper[:20]):
            cedula = nc
            break
    
    if not cedula:
        for nk, nc in name_to_cedula.items():
            if nk in nombre_upper or nombre_upper in nk:
                cedula = nc
                break
    
    if not cedula:
        for emp in all_employees.values():
            if emp['nombre'].split()[-1].upper() in nombre_upper:
                cedula = emp['cedula']
                break
    
    staff_id = staff_id_map.get(cedula) if cedula else None
    if not staff_id:
        print(f"  ⚠️ No match HE para: {rec['nombre']}")
        continue
    
    sede = rec['sede']
    periodo_id = periodo_by_sede.get(sede)
    if not periodo_id: continue
    
    he_records_db.append({
        'periodo_id': periodo_id,
        'staff_id': staff_id,
        'sede': sede,
        **{k: v for k, v in rec.items() if k != 'nombre' and k != 'sede'},
    })

print(f"  HE/Recargos: {len(he_records_db)} registros")
for i in range(0, len(he_records_db), batch_size):
    batch = he_records_db[i:i+batch_size]
    result = api_call('insert_he_recargos', batch)
    print(f"  Batch {i//batch_size+1}: {result}")

# ═══════════════════════════════════════════════════════
# 9. INSERT NOVEDADES
# ═══════════════════════════════════════════════════════
print("\n📋 Insertando novedades...")
novedades_db = []
for nov in novedades:
    staff_id = staff_id_map.get(nov['cedula'])
    if not staff_id:
        print(f"  ⚠️ No staff_id para novedad de {nov['cedula']}")
        continue
    sede = nov['sede']
    periodo_id = periodo_by_sede.get(sede)
    if not periodo_id: continue
    
    novedades_db.append({
        'periodo_id': periodo_id,
        'staff_id': staff_id,
        'sede': sede,
        'tipo': nov['tipo'],
        'observacion': nov.get('observacion', ''),
        'dias': nov.get('dias'),
    })

print(f"  Novedades: {len(novedades_db)} registros")
result = api_call('insert_novedades', novedades_db)
print(f"  Resultado: {result}")

# ═══════════════════════════════════════════════════════
# 10. INSERT PROVISIONES
# ═══════════════════════════════════════════════════════
print("\n💰 Insertando provisiones...")
provisiones_db = []
for prov in provisiones:
    cedula = prov['cedula']
    staff_id = staff_id_map.get(cedula)
    
    if not staff_id:
        nombre_upper = prov['nombre'].upper().strip()
        for emp in all_employees.values():
            if emp['nombre'].upper().strip() == nombre_upper or emp['nombre'].split()[-1].upper() in nombre_upper:
                staff_id = staff_id_map.get(emp['cedula'])
                break
    
    if not staff_id:
        print(f"  ⚠️ No staff_id para prov de {cedula} ({prov['nombre']})")
        continue
    
    sede = prov['sede']
    periodo_id = periodo_by_sede.get(sede)
    if not periodo_id: continue
    
    total_prov_emp = (prov['pension_empleador'] + prov['arl_empleador'] + prov['caja_empleador'] +
                      prov['cesantias_empleador'] + prov['prima_empleador'] + prov['vacaciones_empleador'] +
                      prov['intereses_cesantias_empleador'])
    
    provisiones_db.append({
        'periodo_id': periodo_id,
        'staff_id': staff_id,
        'sede': sede,
        'provisiones_salud': prov['provisiones_salud'],
        'provisiones_sociales': prov['provisiones_sociales'],
        'base_vacaciones': prov['base_vacaciones'],
        'salud_empleado': prov['salud_empleado'],
        'pension_empleado': prov['pension_empleado'],
        'pension_empleador': prov['pension_empleador'],
        'arl_empleador': prov['arl_empleador'],
        'caja_empleador': prov['caja_empleador'],
        'cesantias_empleador': prov['cesantias_empleador'],
        'prima_empleador': prov['prima_empleador'],
        'vacaciones_empleador': prov['vacaciones_empleador'],
        'intereses_cesantias_empleador': prov['intereses_cesantias_empleador'],
        'total_provision_empleador': round(total_prov_emp, 2),
    })

print(f"  Provisiones: {len(provisiones_db)} registros")
for i in range(0, len(provisiones_db), batch_size):
    batch = provisiones_db[i:i+batch_size]
    result = api_call('insert_provisiones', batch)
    print(f"  Batch {i//batch_size+1}: {result}")

# ═══════════════════════════════════════════════════════
# 11. INSERT PROPINAS
# ═══════════════════════════════════════════════════════
print("\n💵 Insertando propinas...")
propinas_data = [
    {'sede': 'C75', 'total_propinas_ventas': 60789692, 'prometidos_100_pct': 0, 'propina_para_rep': 60789692, 'dias_laborados_total': 1059, 'valor_dia_propina': 57402.92},
    {'sede': 'C85', 'total_propinas_ventas': 2827742, 'prometidos_100_pct': 0, 'propina_para_rep': 2827742, 'dias_laborados_total': 105, 'valor_dia_propina': 26930.88},
    {'sede': 'KINDER', 'total_propinas_ventas': 1969675, 'prometidos_100_pct': 0, 'propina_para_rep': 1969675, 'dias_laborados_total': 90, 'valor_dia_propina': 21885.28},
]

propinas_records = []
for p in propinas_data:
    sede = p['sede']
    periodo_id = periodo_by_sede.get(sede)
    if not periodo_id: continue
    propinas_records.append({**p, 'periodo_id': periodo_id})

result = api_call('insert_propinas', propinas_records)
print(f"  Resultado: {result}")

# ═══════════════════════════════════════════════════════
# 12. UPDATE PERIODOS WITH TOTALS
# ═══════════════════════════════════════════════════════
print("\n📊 Actualizando totales de periodos...")
totals_by_sede = {}
for emp in all_employees.values():
    sede = emp['sede']
    if sede not in totals_by_sede:
        totals_by_sede[sede] = {'devengado': 0, 'deducciones': 0, 'neto': 0}
    totals_by_sede[sede]['devengado'] += emp['total_devengado']
    totals_by_sede[sede]['deducciones'] += emp['total_deducciones']
    totals_by_sede[sede]['neto'] += emp['neto_a_pagar']

for sede, totals in totals_by_sede.items():
    periodo_id = periodo_by_sede.get(sede)
    if periodo_id:
        result = api_call('update_periodo', {
            'id': periodo_id,
            'total_devengado': round(totals['devengado'], 2),
            'total_deducciones': round(totals['deducciones'], 2),
            'total_neto': round(totals['neto'], 2),
        })
        print(f"  {sede}: Dev=${totals['devengado']:,.0f} Ded=${totals['deducciones']:,.0f} Neto=${totals['neto']:,.0f} → {result}")

# ═══════════════════════════════════════════════════════
# RESUMEN
# ═══════════════════════════════════════════════════════
print("\n" + "=" * 60)
print("✅ IMPORTACIÓN COMPLETADA")
print("=" * 60)
print(f"  Empleados upserted: {len(empleados_to_upsert)}")
print(f"  Detalle nómina: {len(detalle_records)} registros")
print(f"  HE/Recargos: {len(he_records_db)} registros")
print(f"  Novedades: {len(novedades_db)} registros")
print(f"  Provisiones: {len(provisiones_db)} registros")
print(f"  Propinas: {len(propinas_records)} registros")
print("=" * 60)